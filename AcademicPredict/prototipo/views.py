# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Max, Min
from django.utils import timezone
from django.urls import reverse
from datetime import datetime
import traceback
import json

# Imports de utilidades (ahora centralizadas)
from .utils.permissions import ( puede_ver_anomalias)
from .utils.notifications import (enviar_notificacion_derivacion, enviar_notificacion_cambio_estado)
from .utils.helpers import (determinar_nivel_criticidad)

# Imports de servicios
from .services.import_service import ImportService
from .services.reports_service import ReportsService

# Imports de modelos y formularios
from .models import (DeteccionAnomalia, CriterioAnomalia, Derivacion, Estudiante, Carrera, EjecucionAnalisis,InstanciaApoyo, Asignatura, RegistroAcademico,Usuario)
from .forms import (CriterioAnomaliaForm, DerivacionForm, FiltroAnomaliasForm, ImportarDatosForm)
from .ML import ejecutar_deteccion_anomalias

@login_required
def dashboard(request):
    """Dashboard CORREGIDO con asignaturas críticas para todos los roles."""
    context = {}
    
    try:
        print(f"🏠 Dashboard cargando para usuario: {request.user.username} ({request.user.rol})")
        
        # Obtener datos base
        estudiantes = Estudiante.objects.filter(activo=True)
        anomalias = DeteccionAnomalia.objects.all()
        carrera = None
        
        # Filtrar por rol
        if request.user.rol == 'coordinador_carrera':
            try:
                carrera = Carrera.objects.get(coordinador=request.user)
                estudiantes = estudiantes.filter(carrera=carrera)
                anomalias = anomalias.filter(estudiante__carrera=carrera)
                print(f"👨‍🎓 Filtrando por carrera: {carrera.nombre}")
            except Carrera.DoesNotExist:
                messages.warning(request, "Tu usuario no tiene carrera asignada.")
        
        # Calcular métricas FRESCAS
        total_estudiantes = estudiantes.count()
        total_anomalias = anomalias.count()
        
        # Anomalías activas (no resueltas)
        anomalias_activas = anomalias.filter(
            estado__in=['detectado', 'en_revision', 'intervencion_activa']
        ).count()
        
        # Casos críticos (prioridad alta)
        anomalias_criticas = anomalias.filter(
            prioridad__gte=4,
            estado__in=['detectado', 'en_revision', 'intervencion_activa']
        ).count()
        
        # Derivaciones pendientes
        derivaciones_pendientes = Derivacion.objects.filter(
            estado__in=['pendiente', 'enviada']
        )
        
        if request.user.rol == 'coordinador_carrera' and carrera:
            # Filtrar derivaciones por carrera
            derivaciones_pendientes = derivaciones_pendientes.filter(
                deteccion_anomalia__estudiante__carrera=carrera
            )
        
        total_derivaciones_pendientes = derivaciones_pendientes.count()
        
        # Tasa de anomalías
        tasa_anomalias = round((total_anomalias / total_estudiantes * 100), 2) if total_estudiantes > 0 else 0
        
        # Últimas detecciones (5 más recientes)
        ultimas_anomalias = anomalias.filter(
            estado='detectado'
        ).select_related('estudiante', 'criterio_usado').order_by('-fecha_deteccion')[:5]
        
        # CALCULAR ASIGNATURAS CRÍTICAS para todos los roles
        asignaturas_criticas = []
        
        try:
            if request.user.rol == 'coordinador_carrera' and carrera:
                # Para coordinadores de carrera: solo su carrera
                asignaturas_base = Asignatura.objects.filter(carrera=carrera)
                print(f"📚 Analizando {asignaturas_base.count()} asignaturas de {carrera.nombre}")
            elif request.user.rol in ['coordinador_cpa', 'analista_cpa','admin']:
                # Para CPA: todas las carreras
                asignaturas_base = Asignatura.objects.all()
                print(f"📚 Analizando {asignaturas_base.count()} asignaturas totales")
            else:
                asignaturas_base = Asignatura.objects.none()
            
            for asignatura in asignaturas_base[:20]:  # Limitar para performance
                # Obtener estudiantes de esta asignatura
                registros = RegistroAcademico.objects.filter(
                    asignatura=asignatura,
                    estudiante__activo=True
                )
                
                if registros.exists():
                    # Estudiantes únicos en la asignatura
                    estudiantes_ids = list(registros.values_list('estudiante_id', flat=True).distinct())
                    total_estudiantes_asignatura = len(estudiantes_ids)
                    
                    if total_estudiantes_asignatura > 0:
                        # Contar anomalías activas para estos estudiantes
                        anomalias_asignatura = DeteccionAnomalia.objects.filter(
                            estudiante_id__in=estudiantes_ids,
                            estado__in=['detectado', 'en_revision', 'intervencion_activa']
                        ).count()
                        
                        # Calcular porcentaje
                        porcentaje_anomalias = round((anomalias_asignatura / total_estudiantes_asignatura) * 100, 2)
                        
                        print(f"   📊 {asignatura.nombre}: {anomalias_asignatura}/{total_estudiantes_asignatura} = {porcentaje_anomalias}%")
                        
                        # Solo incluir si es crítica (≥15% anomalías)
                        if porcentaje_anomalias >= 15.0:
                            criticidad_label = 'media'
                            if porcentaje_anomalias >= 30.0:
                                criticidad_label = 'alta'
                            
                            asignaturas_criticas.append({
                                'asignatura': asignatura,
                                'porcentaje_anomalias': porcentaje_anomalias,
                                'total_estudiantes': total_estudiantes_asignatura,
                                'estudiantes_anomalos': anomalias_asignatura,
                                'nivel_criticidad': criticidad_label
                            })
            
            # Ordenar por porcentaje de anomalías (más críticas primero)
            asignaturas_criticas.sort(key=lambda x: x['porcentaje_anomalias'], reverse=True)
            
            # Limitar a top 10 para el dashboard
            asignaturas_criticas = asignaturas_criticas[:10]
            
            print(f"🚨 Total asignaturas críticas encontradas: {len(asignaturas_criticas)}")
            
        except Exception as e:
            print(f"⚠️ Error calculando asignaturas críticas: {e}")
            import traceback
            traceback.print_exc()
            asignaturas_criticas = []
        
        # Preparar contexto
        context.update({
            'total_estudiantes': total_estudiantes,
            'anomalias_activas': anomalias_activas,
            'anomalias_criticas': anomalias_criticas,
            'derivaciones_pendientes': total_derivaciones_pendientes,
            'ultimas_anomalias': ultimas_anomalias,
            'asignaturas_criticas': asignaturas_criticas,
            'ultima_actualizacion': timezone.now(),
            'carrera': carrera,
            'usuario_rol': request.user.rol,
        })
        
        print(f"📊 Dashboard cargado:")
        print(f"   Total estudiantes: {total_estudiantes}")
        print(f"   Total anomalías: {total_anomalias}")
        print(f"   Anomalías activas: {anomalias_activas}")
        print(f"   Casos críticos: {anomalias_criticas}")
        print(f"   Asignaturas críticas: {len(asignaturas_criticas)}")
        
    except Exception as e:
        print(f"❌ Error en dashboard: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f'Error cargando dashboard: {str(e)}')
        
        # Valores por defecto en caso de error
        context.update({
            'total_estudiantes': 0,
            'total_anomalias': 0,
            'anomalias_activas': 0,
            'anomalias_criticas': 0,
            'derivaciones_pendientes': 0,
            'ultimas_anomalias': [],
            'asignaturas_criticas': [],
            'error': True
        })
    
    return render(request, 'anomalias/dashboard.html', context)

@login_required
def exportar_reporte_anomalias(request):
    """
    Exporta reporte de anomalías a Excel
    
    🎓 APRENDIZAJE: Usa el servicio de reportes
    """
    return ReportsService.exportar_anomalias_completo(request, formato='excel')

@login_required
def asignaturas_criticas(request):
    """
    Vista FINAL para asignaturas críticas - FUNCIONA según el debug
    """
    try:
        print(f"\n🏫 === ASIGNATURAS CRÍTICAS FINAL ===")
        print(f"Usuario: {request.user.username} ({request.user.rol})")
        
        carrera = None
        asignaturas_query = Asignatura.objects.all()
        
        # Filtrar según el rol del usuario
        if request.user.rol == 'coordinador_carrera':
            try:
                carrera = Carrera.objects.get(coordinador=request.user)
                asignaturas_query = asignaturas_query.filter(carrera=carrera)
                print(f"👨‍🎓 Coordinador de carrera - Filtrando por: {carrera.nombre}")
            except Carrera.DoesNotExist:
                messages.error(request, "Tu usuario no tiene una carrera asignada.")
                return redirect('dashboard')
        
        elif request.user.rol in ['coordinador_cpa', 'analista_cpa','admin']:
            print(f"👑 {request.user.rol} - Acceso a todas las carreras")
            carrera = None
        
        else:
            messages.error(request, "No tienes permisos para acceder a esta sección.")
            return redirect('dashboard')
        
        # Obtener asignaturas
        asignaturas = asignaturas_query.select_related('carrera')
        print(f"📚 Total asignaturas encontradas: {asignaturas.count()}")
        
        if not asignaturas.exists():
            print("⚠️ No hay asignaturas en el sistema")
            messages.warning(request, "No hay asignaturas registradas en el sistema.")
            
            return render(request, 'anomalias/asignaturas_criticas.html', {
                'asignaturas_criticas': [],
                'total_asignaturas': 0,
                'total_criticas': 0,
                'promedio_anomalias_carrera': 0,
                'carrera': carrera,
                'umbral_criticidad': 15.0,
                'mostrar_todas_carreras': request.user.rol in ['coordinador_cpa', 'analista_cpa'],
                'debug_info': 'No hay asignaturas disponibles'
            })
        
        # Analizar cada asignatura
        asignaturas_criticas = []
        total_asignaturas = 0
        suma_porcentajes = 0
        
        print(f"\n🔍 Analizando asignaturas...")
        
        for asignatura in asignaturas:
            print(f"\n📖 Procesando: {asignatura.nombre}")
            
            # Obtener registros académicos para esta asignatura
            registros = RegistroAcademico.objects.filter(
                asignatura=asignatura,
                estudiante__activo=True
            )
            
            if not registros.exists():
                print(f"   ⚠️ Sin registros para {asignatura.nombre}")
                continue
            
            # Obtener estudiantes únicos
            estudiantes_ids = list(registros.values_list('estudiante_id', flat=True).distinct())
            total_estudiantes_asignatura = len(estudiantes_ids)
            
            print(f"   👥 Estudiantes únicos: {total_estudiantes_asignatura}")
            
            if total_estudiantes_asignatura == 0:
                continue
            
            # Buscar anomalías ACTIVAS para estos estudiantes
            anomalias_estudiantes = DeteccionAnomalia.objects.filter(
                estudiante_id__in=estudiantes_ids,
                estado__in=['detectado', 'en_revision', 'intervencion_activa']
            )
            
            anomalias_count = anomalias_estudiantes.count()
            print(f"   ⚠️ Anomalías activas: {anomalias_count}")
            
            # Calcular porcentaje
            porcentaje_anomalias = (anomalias_count / total_estudiantes_asignatura) * 100
            print(f"   📊 Porcentaje de anomalías: {porcentaje_anomalias:.2f}%")
            
            # Determinar nivel de criticidad
            nivel_criticidad = determinar_nivel_criticidad(porcentaje_anomalias)
            
            # Agregar a la lista si es crítica (≥15% de anomalías)
            if porcentaje_anomalias >= 15.0:
                print(f"   🔴 CRÍTICA: {asignatura.nombre} - {porcentaje_anomalias:.2f}%")
                
                asignaturas_criticas.append({
                    'asignatura': asignatura,
                    'total_estudiantes': total_estudiantes_asignatura,
                    'total_anomalias': anomalias_count,
                    'porcentaje_anomalias': round(porcentaje_anomalias, 2),
                    'nivel_criticidad': nivel_criticidad,
                    'carrera_nombre': asignatura.carrera.nombre if asignatura.carrera else 'Sin carrera'
                })
            else:
                print(f"   ✅ Normal: {asignatura.nombre} - {porcentaje_anomalias:.2f}%")
            
            total_asignaturas += 1
            suma_porcentajes += porcentaje_anomalias
        
        # Ordenar por porcentaje de anomalías (de mayor a menor)
        asignaturas_criticas.sort(key=lambda x: x['porcentaje_anomalias'], reverse=True)
        
        # Calcular promedio
        promedio_anomalias_carrera = suma_porcentajes / total_asignaturas if total_asignaturas > 0 else 0
        
        print(f"\n📊 === RESULTADOS FINALES ===")
        print(f"   Total asignaturas analizadas: {total_asignaturas}")
        print(f"   Asignaturas críticas encontradas: {len(asignaturas_criticas)}")
        print(f"   Promedio de anomalías: {promedio_anomalias_carrera:.2f}%")
        
        if len(asignaturas_criticas) > 0:
            print(f"   🔴 ¡Se encontraron {len(asignaturas_criticas)} asignaturas críticas!")
            for critica in asignaturas_criticas[:5]:  # Mostrar solo las primeras 5
                print(f"      - {critica['asignatura'].nombre}: {critica['porcentaje_anomalias']}%")
        
        context = {
            'asignaturas_criticas': asignaturas_criticas,
            'total_asignaturas': total_asignaturas,
            'total_criticas': len(asignaturas_criticas),
            'promedio_anomalias_carrera': round(promedio_anomalias_carrera, 2),
            'carrera': carrera,
            'umbral_criticidad': 15.0,
            'mostrar_todas_carreras': request.user.rol in ['coordinador_cpa', 'analista_cpa'],
            'usuario_rol': request.user.rol,
            'debug_info': f"Analizadas {total_asignaturas} asignaturas - {len(asignaturas_criticas)} críticas encontradas"
        }
        
        return render(request, 'anomalias/asignaturas_criticas.html', context)
        
    except Exception as e:
        print(f"❌ Error en asignaturas_criticas: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f'Error analizando asignaturas críticas: {str(e)}')
        return render(request, 'anomalias/asignaturas_criticas.html', {
            'asignaturas_criticas': [],
            'total_asignaturas': 0,
            'total_criticas': 0,
            'promedio_anomalias_carrera': 0,
            'carrera': None,
            'umbral_criticidad': 15.0,
            'error': True,
            'error_message': str(e)
        })


@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa','admin'])
def gestionar_derivaciones(request):
    """Vista mejorada para gestionar derivaciones."""
    # Queryset base
    derivaciones = Derivacion.objects.select_related(
        'deteccion_anomalia__estudiante',
        'deteccion_anomalia__estudiante__carrera',
        'instancia_apoyo',
        'derivado_por'
    ).order_by('-fecha_derivacion')
    
    # Aplicar filtros
    estado = request.GET.get('estado')
    if estado:
        derivaciones = derivaciones.filter(estado=estado)
    
    instancia = request.GET.get('instancia')
    if instancia:
        derivaciones = derivaciones.filter(instancia_apoyo_id=instancia)
    
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        derivaciones = derivaciones.filter(fecha_derivacion__date__gte=fecha_desde)
    
    busqueda = request.GET.get('busqueda')
    if busqueda:
        derivaciones = derivaciones.filter(
            Q(deteccion_anomalia__estudiante__nombre__icontains=busqueda) |
            Q(deteccion_anomalia__estudiante__id_estudiante__icontains=busqueda)
        )
    
    # Estadísticas rápidas
    total_derivaciones = derivaciones.count()
    derivaciones_pendientes = derivaciones.filter(estado='pendiente').count()
    derivaciones_proceso = derivaciones.filter(estado='en_proceso').count()
    derivaciones_completadas = derivaciones.filter(estado='completada').count()
    
    # Paginación
    paginator = Paginator(derivaciones, 15)
    page = request.GET.get('page')
    derivaciones_paginadas = paginator.get_page(page)
    
    context = {
        'derivaciones': derivaciones_paginadas,
        'derivaciones_pendientes': derivaciones_pendientes,
        'derivaciones_proceso': derivaciones_proceso,
        'derivaciones_completadas': derivaciones_completadas,
        'total_derivaciones': total_derivaciones,
        'estados_derivacion': Derivacion.ESTADOS_DERIVACION,
        'instancias_apoyo': InstanciaApoyo.objects.filter(activo=True),
    }
    
    return render(request, 'anomalias/gestionar_derivaciones.html', context)

# Vista para gestión masiva de anomalías
@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa','admin'])
def gestion_masiva_anomalias(request):
    """
    🔧 FUNCIÓN MEJORADA: Gestión masiva de anomalías
    """
    if request.method != 'POST':
        messages.warning(request, 'Método no permitido para gestión masiva.')
        return redirect('listado_anomalias')
    
    try:
        # Obtener parámetros del formulario
        action = request.POST.get('action')
        anomalia_ids = request.POST.getlist('anomalias_seleccionadas')
        
        print(f"🔍 Gestión masiva:")
        print(f"   Action: {action}")
        print(f"   IDs: {anomalia_ids}")
        print(f"   Usuario: {request.user.username}")
        
        # CASO ESPECIAL: exportar_filtrados (sin IDs específicos)
        if action == 'exportar_filtrados':
            print("📊 Redirigiendo a exportación filtrada...")
            # Mantener los parámetros GET para aplicar los mismos filtros
            query_params = request.GET.urlencode()
            redirect_url = f"{reverse('exportar_todas_anomalias')}?{query_params}"
            return redirect(redirect_url)
        
        # Para las demás acciones, validar que se seleccionaron anomalías
        if not anomalia_ids:
            messages.error(request, 'No se seleccionaron anomalías.')
            return redirect('listado_anomalias')
        
        # Convertir IDs a enteros
        try:
            anomalia_ids = [int(id) for id in anomalia_ids]
        except ValueError:
            messages.error(request, 'IDs de anomalías inválidos.')
            return redirect('listado_anomalias')
        
        # Obtener anomalías
        anomalias = DeteccionAnomalia.objects.filter(id__in=anomalia_ids)
        
        # Filtrar por permisos del usuario
        if request.user.rol == 'coordinador_carrera':
            try:
                carrera = Carrera.objects.get(coordinador=request.user)
                anomalias = anomalias.filter(estudiante__carrera=carrera)
            except Carrera.DoesNotExist:
                messages.error(request, 'No tienes permisos para esta acción.')
                return redirect('listado_anomalias')
        
        if not anomalias.exists():
            messages.error(request, 'No se encontraron anomalías válidas.')
            return redirect('listado_anomalias')
        
        # Ejecutar acción según el tipo
        if action == 'cambiar_estado':
            nuevo_estado = request.POST.get('nuevo_estado')
            if nuevo_estado in dict(DeteccionAnomalia.ESTADOS):
                count = anomalias.update(
                    estado=nuevo_estado,
                    revisado_por=request.user,
                    fecha_ultima_actualizacion=timezone.now()
                )
                messages.success(request, f'Se actualizó el estado de {count} anomalías a "{dict(DeteccionAnomalia.ESTADOS)[nuevo_estado]}".')
            else:
                messages.error(request, 'Estado inválido.')

        elif action == 'exportar':
            # Exportar solo las anomalías seleccionadas
            return generar_reporte_anomalias_seleccionadas(anomalias, request)

        elif action == 'derivar_masivo':
            # Derivar masivamente
            instancia_id = request.POST.get('instancia_apoyo')
            motivo = request.POST.get('motivo', '')
            prioridad = request.POST.get('prioridad', 3)

            try:
                instancia = InstanciaApoyo.objects.get(id=instancia_id, activo=True)
                count_derivadas = 0

                for anomalia in anomalias:
                    # Solo derivar si puede ser derivada
                    if anomalia.puede_ser_derivada():
                        Derivacion.objects.create(
                            deteccion_anomalia=anomalia,
                            instancia_apoyo=instancia,
                            motivo=motivo,
                            prioridad=int(prioridad),
                            derivado_por=request.user,
                            estado='pendiente'
                        )
                        # Actualizar estado de la anomalía
                        anomalia.estado = 'intervencion_activa'
                        anomalia.revisado_por = request.user
                        anomalia.fecha_ultima_actualizacion = timezone.now()
                        anomalia.save()
                        count_derivadas += 1

                if count_derivadas > 0:
                    messages.success(request, f'Se derivaron {count_derivadas} anomalías a "{instancia.nombre}".')
                else:
                    messages.warning(request, 'No se pudo derivar ninguna anomalía. Verifica que estén en estado válido.')

            except InstanciaApoyo.DoesNotExist:
                messages.error(request, 'Instancia de apoyo no encontrada.')
            except ValueError as e:
                messages.error(request, f'Error en los datos: {str(e)}')

        else:
            messages.error(request, f'Acción no válida: {action}.')
        
        return redirect('listado_anomalias')
        
    except Exception as e:
        print(f"❌ Error en gestión masiva: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f'Error en gestión masiva: {str(e)}')
        return redirect('listado_anomalias')

# Vista para actualizar estado de derivación CORREGIDA
@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa','admin'])
def actualizar_estado_derivacion(request, derivacion_id):
    """
    🔧 FUNCIÓN CORREGIDA: Actualizar estado de derivación
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        derivacion = get_object_or_404(Derivacion, id=derivacion_id)
        
        nuevo_estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')
        
        if nuevo_estado not in dict(Derivacion.ESTADOS_DERIVACION):
            return JsonResponse({'error': 'Estado inválido'}, status=400)
        
        # Actualizar derivación
        estado_anterior = derivacion.get_estado_display()
        derivacion.estado = nuevo_estado
        
        # 🔧 AGREGAR OBSERVACIONES AL CAMPO CORRECTO
        if observaciones:
            timestamp = timezone.now().strftime('%d/%m/%Y %H:%M')
            usuario = request.user.get_full_name() or request.user.username
            nueva_observacion = f"[{timestamp}] {usuario}: {observaciones}"
            
            # Verificar si existe campo observaciones_seguimiento
            if hasattr(derivacion, 'observaciones_seguimiento'):
                # Si existe el campo, usarlo
                if derivacion.observaciones_seguimiento:
                    derivacion.observaciones_seguimiento += f"\n\n{nueva_observacion}"
                else:
                    derivacion.observaciones_seguimiento = nueva_observacion
            else:
                # Si no existe, usar observaciones_derivacion
                if derivacion.observaciones_derivacion:
                    derivacion.observaciones_derivacion += f"\n\n{nueva_observacion}"
                else:
                    derivacion.observaciones_derivacion = f"SEGUIMIENTO:\n{nueva_observacion}"
        
        derivacion.save()
        
        # Si la derivación se completa, actualizar la anomalía
        if nuevo_estado == 'completada':
            derivacion.deteccion_anomalia.estado = 'resuelto'
            derivacion.deteccion_anomalia.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Estado actualizado de "{estado_anterior}" a "{derivacion.get_estado_display()}"',
            'nuevo_estado': nuevo_estado,
            'nuevo_estado_display': derivacion.get_estado_display()
        })
        
    except Exception as e:
        print(f"❌ Error actualizando derivación: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def verificar_sistema(request):
    """Vista para verificación rápida del sistema."""
    
    # Estadísticas básicas
    stats = {
        'estudiantes_activos': Estudiante.objects.filter(activo=True).count(),
        'registros_academicos': RegistroAcademico.objects.count(),
        'criterios_activos': CriterioAnomalia.objects.filter(activo=True).count(),
        'anomalias_total': DeteccionAnomalia.objects.count(),
        'anomalias_activas': DeteccionAnomalia.objects.filter(
            estado__in=['detectado', 'en_revision', 'intervencion_activa']
        ).count(),
        'ejecuciones_exitosas': EjecucionAnalisis.objects.filter(exitoso=True).count(),
        'ultima_ejecucion': EjecucionAnalisis.objects.order_by('-fecha_ejecucion').first()
    }
    
    # Problemas detectados
    problemas = []
    
    if stats['estudiantes_activos'] < 10:
        problemas.append('Muy pocos estudiantes activos (< 10)')
    
    if stats['registros_academicos'] < 30:
        problemas.append('Muy pocos registros académicos (< 30)')
    
    if stats['criterios_activos'] == 0:
        problemas.append('No hay criterios activos')
    
    if stats['anomalias_total'] == 0 and stats['ejecuciones_exitosas'] > 0:
        problemas.append('Hay ejecuciones exitosas pero no hay anomalías guardadas')
    
    # Distribución por estudiante
    if stats['estudiantes_activos'] > 0 and stats['registros_academicos'] > 0:
        registros_por_estudiante = stats['registros_academicos'] / stats['estudiantes_activos']
        if registros_por_estudiante < 3:
            problemas.append(f'Pocos registros por estudiante ({registros_por_estudiante:.1f} < 3)')
    
    context = {
        'stats': stats,
        'problemas': problemas,
        'estado_general': 'OK' if not problemas else 'PROBLEMAS DETECTADOS'
    }
    
    return render(request, 'anomalias/verificar_sistema.html', context)

# 🔧 VERIFICACIÓN RÁPIDA: Función para confirmar el nombre correcto
def verificar_campo_ingreso():
    """
    🔍 Verificación rápida del campo de año de ingreso
    """
    try:
        estudiante = Estudiante.objects.first()
        if estudiante:
            print("🔍 Verificando campos de año de ingreso:")
            
            campos_posibles = ['ingreso_año', 'ingreso_ano', 'año_ingreso', 'ano_ingreso']
            
            for campo in campos_posibles:
                if hasattr(estudiante, campo):
                    valor = getattr(estudiante, campo)
                    print(f"   ✅ {campo}: {valor}")
                else:
                    print(f"   ❌ {campo}: NO EXISTE")
        else:
            print("❌ No hay estudiantes en la base de datos")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

@login_required
def ayuda_documentacion(request):
    """Vista para mostrar ayuda y documentación"""
    return render(request, 'anomalias/ayuda_documentacion.html')

def generar_reporte_anomalias_seleccionadas(anomalias_queryset, request):
    """
    Genera un reporte Excel de las anomalías seleccionadas
    
    🎓 EDUCATIVO: Esta función crea un archivo Excel con:
    - Datos del estudiante
    - Métricas de la anomalía
    - Estado actual
    - Derivaciones asociadas
    
    Args:
        anomalias_queryset: QuerySet de DeteccionAnomalia
        request: HttpRequest
        
    Returns:
        HttpResponse con archivo Excel
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.utils import timezone
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Anomalías Seleccionadas"
    
    # ================================================================
    # ENCABEZADOS
    # ================================================================
    headers = [
        'ID', 'Estudiante', 'ID Estudiante', 'Carrera',
        'Tipo Anomalía', 'Estado', 'Prioridad',
        'Promedio General', 'Asistencia %', 'Uso Plataforma %',
        'Score Anomalía', 'Confianza %', 
        'Fecha Detección', 'Revisado Por',
        'Tiene Derivación', 'Observaciones'
    ]
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # ================================================================
    # DATOS
    # ================================================================
    anomalias = anomalias_queryset.select_related(
        'estudiante',
        'estudiante__carrera',
        'revisado_por'
    ).prefetch_related('derivacion_set')
    
    for row_num, anomalia in enumerate(anomalias, 2):
        # Verificar si tiene derivación
        tiene_derivacion = anomalia.derivacion_set.exists()
        
        datos = [
            anomalia.id,
            anomalia.estudiante.nombre,
            anomalia.estudiante.id_estudiante,
            anomalia.estudiante.carrera.nombre if anomalia.estudiante.carrera else 'N/A',
            anomalia.get_tipo_anomalia_display(),
            anomalia.get_estado_display(),
            anomalia.prioridad,
            round(anomalia.promedio_general, 2),
            round(anomalia.asistencia_promedio, 1),
            round(anomalia.uso_plataforma_promedio, 1),
            round(anomalia.score_anomalia, 4),
            round(anomalia.confianza, 1),
            anomalia.fecha_deteccion.strftime('%Y-%m-%d %H:%M'),
            anomalia.revisado_por.get_full_name() if anomalia.revisado_por else 'Sin asignar',
            'Sí' if tiene_derivacion else 'No',
            anomalia.observaciones[:100] if anomalia.observaciones else ''
        ]
        
        for col_num, valor in enumerate(datos, 1):
            ws.cell(row=row_num, column=col_num, value=valor)
    
    # ================================================================
    # AJUSTAR ANCHOS DE COLUMNA
    # ================================================================
    column_widths = {
        'A': 8,   # ID
        'B': 25,  # Estudiante
        'C': 15,  # ID Estudiante
        'D': 30,  # Carrera
        'E': 20,  # Tipo
        'F': 15,  # Estado
        'G': 10,  # Prioridad
        'H': 12,  # Promedio
        'I': 12,  # Asistencia
        'J': 15,  # Uso Plataforma
        'K': 12,  # Score
        'L': 12,  # Confianza
        'M': 18,  # Fecha
        'N': 20,  # Revisado Por
        'O': 15,  # Tiene Derivación
        'P': 40,  # Observaciones
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # ================================================================
    # HOJA DE RESUMEN
    # ================================================================
    ws_resumen = wb.create_sheet(title="Resumen")
    
    # Estadísticas
    total = anomalias.count()
    por_estado = anomalias.values('estado').annotate(
        count=Count('id')
    ).order_by('-count')
    
    por_tipo = anomalias.values('tipo_anomalia').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Escribir resumen
    ws_resumen['A1'] = 'RESUMEN DE ANOMALÍAS SELECCIONADAS'
    ws_resumen['A1'].font = Font(size=14, bold=True)
    
    ws_resumen['A3'] = f'Total de anomalías: {total}'
    ws_resumen['A4'] = f'Fecha de generación: {timezone.now().strftime("%Y-%m-%d %H:%M")}'
    ws_resumen['A5'] = f'Generado por: {request.user.get_full_name()}'
    
    # Distribución por estado
    ws_resumen['A7'] = 'DISTRIBUCIÓN POR ESTADO'
    ws_resumen['A7'].font = Font(bold=True)
    row = 8
    for item in por_estado:
        ws_resumen[f'A{row}'] = item['estado']
        ws_resumen[f'B{row}'] = item['count']
        ws_resumen[f'C{row}'] = f"{(item['count']/total)*100:.1f}%"
        row += 1
    
    # Distribución por tipo
    ws_resumen[f'A{row+1}'] = 'DISTRIBUCIÓN POR TIPO'
    ws_resumen[f'A{row+1}'].font = Font(bold=True)
    row += 2
    for item in por_tipo:
        ws_resumen[f'A{row}'] = item['tipo_anomalia']
        ws_resumen[f'B{row}'] = item['count']
        ws_resumen[f'C{row}'] = f"{(item['count']/total)*100:.1f}%"
        row += 1
    
    # ================================================================
    # PREPARAR RESPUESTA HTTP
    # ================================================================
    from io import BytesIO
    
    # Guardar en memoria
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Crear respuesta
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    filename = f'anomalias_seleccionadas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def perfil_usuario(request):
    """
    👤 FUNCIÓN CORREGIDA: Vista de perfil de usuario
    """
    print(f"👤 Cargando perfil para usuario: {request.user.username}")
    
    try:
        if request.method == 'POST':
            print("📝 Procesando actualización de perfil...")
            
            # Obtener datos del formulario
            nombre = request.POST.get('first_name', '').strip()
            apellido = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            telefono = request.POST.get('telefono', '').strip()
            
            # Validaciones básicas
            errores = []
            
            if not nombre:
                errores.append("El nombre es obligatorio")
            
            if not email:
                errores.append("El email es obligatorio")
            elif not '@' in email:
                errores.append("El email no es válido")
            
            # Verificar si el email ya existe (excepto el usuario actual)
            if Usuario.objects.filter(email=email).exclude(id=request.user.id).exists():
                errores.append("Este email ya está en uso por otro usuario")
            
            if errores:
                for error in errores:
                    messages.error(request, error)
            else:
                # Actualizar datos del usuario
                request.user.first_name = nombre
                request.user.last_name = apellido
                request.user.email = email
                
                # Actualizar teléfono si el campo existe
                if hasattr(request.user, 'telefono'):
                    request.user.telefono = telefono
                
                request.user.save()
                
                messages.success(request, 'Perfil actualizado exitosamente.')
                print(f"✅ Perfil actualizado para {request.user.username}")
                
                return redirect('perfil_usuario')
        
        # Calcular estadísticas del usuario
        stats = {}
        
        # Estadísticas comunes para todos los roles
        if request.user.rol in ['analista_cpa', 'coordinador_cpa']:
            # Derivaciones creadas
            stats['derivaciones_creadas'] = Derivacion.objects.filter(
                derivado_por=request.user
            ).count()
            
            # Anomalías revisadas
            stats['anomalias_revisadas'] = DeteccionAnomalia.objects.filter(
                revisado_por=request.user
            ).count()
            
            # Derivaciones pendientes
            stats['derivaciones_pendientes'] = Derivacion.objects.filter(
                derivado_por=request.user,
                estado__in=['pendiente', 'enviada']
            ).count()
            
            # Anomalías resueltas por el usuario
            stats['anomalias_resueltas'] = DeteccionAnomalia.objects.filter(
                revisado_por=request.user,
                estado='resuelto'
            ).count()
        
        # Estadísticas específicas para coordinador CPA
        if request.user.rol == 'coordinador_cpa':
            stats['criterios_creados'] = CriterioAnomalia.objects.filter(
                creado_por=request.user
            ).count()
            
            stats['analisis_ejecutados'] = EjecucionAnalisis.objects.filter(
                ejecutado_por=request.user
            ).count()
            
            stats['criterios_activos'] = CriterioAnomalia.objects.filter(
                creado_por=request.user,
                activo=True
            ).count()
        
        # Estadísticas para coordinador de carrera
        if request.user.rol == 'coordinador_carrera':
            try:
                carrera = Carrera.objects.get(coordinador=request.user)
                
                stats['estudiantes_carrera'] = Estudiante.objects.filter(
                    carrera=carrera,
                    activo=True
                ).count()
                
                stats['anomalias_carrera'] = DeteccionAnomalia.objects.filter(
                    estudiante__carrera=carrera
                ).count()
                
                stats['asignaturas_carrera'] = Asignatura.objects.filter(
                    carrera=carrera
                ).count()
                
                # Asignaturas críticas
                asignaturas_criticas = 0
                for asignatura in Asignatura.objects.filter(carrera=carrera):
                    registros = RegistroAcademico.objects.filter(asignatura=asignatura)
                    if registros.exists():
                        estudiantes_ids = registros.values_list('estudiante_id', flat=True).distinct()
                        anomalias = DeteccionAnomalia.objects.filter(
                            estudiante_id__in=estudiantes_ids,
                            estado__in=['detectado', 'en_revision', 'intervencion_activa']
                        ).count()
                        
                        if len(estudiantes_ids) > 0:
                            porcentaje = (anomalias / len(estudiantes_ids)) * 100
                            if porcentaje >= 15:
                                asignaturas_criticas += 1
                
                stats['asignaturas_criticas'] = asignaturas_criticas
                stats['carrera_nombre'] = carrera.nombre
                
            except Carrera.DoesNotExist:
                stats['carrera_nombre'] = 'Sin carrera asignada'
                messages.warning(request, "Tu usuario no tiene carrera asignada.")
        
        # Actividad reciente del usuario
        actividad_reciente = []
        
        # Últimas derivaciones
        ultimas_derivaciones = Derivacion.objects.filter(
            derivado_por=request.user
        ).order_by('-fecha_derivacion')[:5]
        
        for derivacion in ultimas_derivaciones:
            actividad_reciente.append({
                'tipo': 'derivacion',
                'descripcion': f'Derivación creada para {derivacion.deteccion_anomalia.estudiante.nombre}',
                'fecha': derivacion.fecha_derivacion,
                'url': reverse('detalle_anomalia', kwargs={'pk': derivacion.deteccion_anomalia.id})
            })
        
        # Últimas anomalías revisadas
        ultimas_revisiones = DeteccionAnomalia.objects.filter(
            revisado_por=request.user
        ).order_by('-fecha_ultima_actualizacion')[:5]
        
        for anomalia in ultimas_revisiones:
            actividad_reciente.append({
                'tipo': 'revision',
                'descripcion': f'Anomalía revisada: {anomalia.estudiante.nombre}',
                'fecha': anomalia.fecha_ultima_actualizacion,
                'url': reverse('detalle_anomalia', kwargs={'pk': anomalia.id})
            })
        
        # Ordenar actividad por fecha
        actividad_reciente.sort(key=lambda x: x['fecha'], reverse=True)
        actividad_reciente = actividad_reciente[:10]  # Top 10
        
        context = {
            'usuario': request.user,
            'stats': stats,
            'actividad_reciente': actividad_reciente,
            'roles_disponibles': Usuario.ROLES,
        }
        
        print(f"📊 Estadísticas calculadas para {request.user.username}: {stats}")
        
        return render(request, 'anomalias/perfil_usuario.html', context)
        
    except Exception as e:
        print(f"❌ Error en perfil_usuario: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f'Error cargando perfil: {str(e)}')
        
        # Contexto mínimo en caso de error
        context = {
            'usuario': request.user,
            'stats': {},
            'actividad_reciente': [],
            'error': True
        }
        
        return render(request, 'anomalias/perfil_usuario.html', context)


@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def importar_datos(request):
    """
    Vista MEJORADA para importar los 3 archivos a la vez,
    usando ImportarDatosForm y el ImportService.
    """
    
    if request.method == 'POST':
        try:
            form = ImportarDatosForm(request.POST, request.FILES)
            if form.is_valid():
                print("📁 Iniciando importación desde web...")
                
                # Este es el diccionario que tu template importar_resultados.html espera
                resultados = {
                    'estudiantes': {'importados': 0, 'errores': [], 'advertencias': []},
                    'asignaturas': {'importados': 0, 'errores': [], 'advertencias': []},
                    'registros': {'importados': 0, 'errores': [], 'advertencias': []}
                }
                
                total_importados = 0
                total_errores = 0
                total_advertencias = 0
                
                # === INICIO DE LA MODIFICACIÓN (usando ImportService) ===
                
                # Procesar archivo de estudiantes
                if form.cleaned_data.get('archivo_estudiantes'):
                    print("👥 Procesando estudiantes...")
                    resultados['estudiantes'] = ImportService.procesar_estudiantes(
                        form.cleaned_data['archivo_estudiantes']
                    )
                    total_importados += resultados['estudiantes']['importados']
                    total_errores += len(resultados['estudiantes']['errores'])
                    total_advertencias += len(resultados['estudiantes']['advertencias'])
                
                # Procesar archivo de asignaturas
                if form.cleaned_data.get('archivo_asignaturas'):
                    print("📚 Procesando asignaturas...")
                    resultados['asignaturas'] = ImportService.procesar_asignaturas(
                        form.cleaned_data['archivo_asignaturas']
                    )
                    total_importados += resultados['asignaturas']['importados']
                    total_errores += len(resultados['asignaturas']['errores'])
                    total_advertencias += len(resultados['asignaturas']['advertencias'])
                
                # Procesar archivo de registros académicos
                if form.cleaned_data.get('archivo_registros'):
                    print("📊 Procesando registros académicos...")
                    resultados['registros'] = ImportService.procesar_registros(
                        form.cleaned_data['archivo_registros']
                    )
                    total_importados += resultados['registros']['importados']
                    total_errores += len(resultados['registros']['errores'])
                    total_advertencias += len(resultados['registros']['advertencias'])
                
                # === FIN DE LA MODIFICACIÓN ===
                
                # Mostrar resultados
                if total_importados > 0:
                    messages.success(
                        request, 
                        f'✅ Importación completada: {total_importados} registros importados/actualizados.'
                    )
                
                if total_errores > 0:
                    messages.warning(
                        request,
                        f'⚠️ Se encontraron {total_errores} errores durante la importación. Revisa los detalles.'
                    )

                if total_advertencias > 0:
                    messages.warning(
                        request,
                        f'⚠️ Se encontraron {total_advertencias} errores durante la importación. Revisa los detalles.'
                    )

                # Renderizar página de resultados (¡esto ahora funcionará!)
                return render(request, 'anomalias/importar_resultados.html', {
                    'resultados': resultados,
                    'total_importados': total_importados,
                    'total_errores': total_errores,
                    'total_advertencias': total_advertencias
                })
                
            else:
                # Si el formulario no es válido, muestra los errores del formulario
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{field}: {error}')
                        
        except Exception as e:
            print(f"❌ Error en importación web: {str(e)}")
            print(traceback.format_exc())
            messages.error(request, f'Error fatal procesando archivos: {str(e)}')
    else:
        form = ImportarDatosForm()
    
    # Obtener estadísticas actuales (para el GET request)
    stats = {
        'total_estudiantes': Estudiante.objects.count(),
        'total_asignaturas': Asignatura.objects.count(), 
        'total_registros': RegistroAcademico.objects.count(),
        'total_carreras': Carrera.objects.count(),
    }
    
    return render(request, 'anomalias/importar_datos.html', {
        'form': form,
        'stats': stats
    })

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def configuracion_criterios(request):
    """
    🔧 CORRECCIÓN: Función que faltaba para configuración de criterios
    
    🎓 EDUCATIVO: Si tu views.py actual tiene una clase en lugar de función,
    esta función será la versión simplificada.
    """
    # Obtener criterios existentes
    criterios = CriterioAnomalia.objects.filter(activo=True).order_by('-fecha_creacion')
    
    # Estadísticas básicas
    estadisticas = {
        'total_criterios': criterios.count(),
        'total_ejecuciones': EjecucionAnalisis.objects.count(),
        'ultima_ejecucion': EjecucionAnalisis.objects.order_by('-fecha_ejecucion').first(),
    }
    
    context = {
        'criterios': criterios,
        'estadisticas': estadisticas,
        'form': CriterioAnomaliaForm(),
    }
    
    return render(request, 'anomalias/configuracion_criterios.html', context)

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def crear_criterio_anomalia(request):
    """
    Crea un nuevo criterio de detección ML
    
    🎓 APRENDIZAJE: Los criterios configuran el algoritmo
    - contamination: % esperado de anomalías
    - n_estimators: Árboles en el Isolation Forest
    """
    
    if request.method == 'POST':
        form = CriterioAnomaliaForm(request.POST)
        if form.is_valid():
            criterio = form.save(commit=False)
            criterio.creado_por = request.user
            criterio.save()
            
            messages.success(request, 'Criterio creado exitosamente')
            return redirect('configuracion_criterios')
    else:
        form = CriterioAnomaliaForm()
    
    return render(request, 'anomalias/crear_criterio.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def detalle_criterio(request, criterio_id):
    """Vista para ver detalles del criterio."""
    criterio = get_object_or_404(CriterioAnomalia, id=criterio_id)
    
    # Obtener estadísticas del criterio
    ejecuciones = EjecucionAnalisis.objects.filter(criterio_usado=criterio).order_by('-fecha_ejecucion')
    total_ejecuciones = ejecuciones.count()
    ejecuciones_exitosas = ejecuciones.filter(exitoso=True).count()
    
    # Anomalías detectadas con este criterio
    anomalias_detectadas = DeteccionAnomalia.objects.filter(criterio_usado=criterio)
    total_anomalias = anomalias_detectadas.count()
    
    # Última ejecución
    ultima_ejecucion = ejecuciones.first()
    
    # Distribución por tipo de anomalía
    anomalias_por_tipo = anomalias_detectadas.values('tipo_anomalia').annotate(
        count=Count('id')
    ).order_by('-count')
    
    context = {
        'criterio': criterio,
        'total_ejecuciones': total_ejecuciones,
        'ejecuciones_exitosas': ejecuciones_exitosas,
        'total_anomalias': total_anomalias,
        'ultima_ejecucion': ultima_ejecucion,
        'anomalias_por_tipo': anomalias_por_tipo,
        'ejecuciones_recientes': ejecuciones[:5],
    }
    
    return render(request, 'anomalias/detalle_criterio.html', context)

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def editar_criterio(request, criterio_id):
    """Vista para editar criterio."""
    criterio = get_object_or_404(CriterioAnomalia, id=criterio_id)
    
    if request.method == 'POST':
        form = CriterioAnomaliaForm(request.POST, instance=criterio)
        if form.is_valid():
            form.save()
            messages.success(request, f'Criterio "{criterio.nombre}" actualizado exitosamente.')
            return redirect('configuracion_criterios')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CriterioAnomaliaForm(instance=criterio)
    
    context = {
        'form': form,
        'criterio': criterio,
        'editando': True,
    }
    
    return render(request, 'anomalias/crear_criterio.html', context)

@login_required
def ejecutar_analisis(request, criterio_id):
    """
    Ejecuta el algoritmo ML con un criterio específico
    
    🎓 APRENDIZAJE: Esta vista conecta la UI con el ML
    - Llama a la función del módulo ML
    - Muestra resultados al usuario
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    criterio = get_object_or_404(CriterioAnomalia, id=criterio_id)
    
    try:
        # Esta línea se ejecuta de forma síncrona y puede tardar
        resultados = ejecutar_deteccion_anomalias(criterio, request.user)
        
        if resultados['exitoso']:
            messages.success(
                request,
                f'Análisis completado: {resultados["anomalias_detectadas"]} anomalías encontradas'
            )
            
            # ✅ CORRECCIÓN: Devolver el resultado final en JSON
            return JsonResponse({
                'success': True,
                'exitoso': True, # Para que coincida con la lógica de finalizarAnalisis
                'anomalias_detectadas': resultados.get('anomalias_detectadas', 0),
                'tiempo_ejecucion': f"{resultados.get('tiempo_ejecucion', 0):.2f}"
            })
        else:
            # Si el ML.py falló
            messages.error(request, f'Error en el análisis: {resultados.get("error", "Error desconocido")}')
            return JsonResponse({
                'success': False, 
                'exitoso': False,
                'error': resultados.get('error', 'Error desconocido')
            }, status=400)
        
    except Exception as e:
        messages.error(request, f'Error fatal en el análisis: {str(e)}')
        traceback.print_exc() # Para ver el error en la consola del servidor
        return JsonResponse({'success': False, 'exitoso': False, 'error': f'Error fatal: {str(e)}'}, status=500)

@login_required
@user_passes_test(lambda u: u.rol in ['admin', 'coordinador_cpa'])
def eliminar_criterio(request, criterio_id):
    """🗑️ Eliminar criterio de detección"""
    try:
        criterio = get_object_or_404(CriterioAnomalia, id=criterio_id)
        
        # Verificar si el criterio tiene anomalías asociadas
        anomalias_asociadas = DeteccionAnomalia.objects.filter(criterio_usado=criterio).count()
        
        if request.method == 'POST':
            confirmar = request.POST.get('confirmar') == 'true'
            
            if confirmar:
                nombre_criterio = criterio.nombre
                
                if anomalias_asociadas > 0:
                    # No eliminar, solo desactivar
                    criterio.activo = False
                    criterio.save()
                    
                    messages.success(
                        request,
                        f'Criterio "{nombre_criterio}" desactivado exitosamente. '
                        f'Se mantiene para preservar el historial de {anomalias_asociadas} anomalías.'
                    )
                else:
                    # Eliminar completamente
                    criterio.delete()
                    
                    messages.success(
                        request,
                        f'Criterio "{nombre_criterio}" eliminado exitosamente.'
                    )
                
                return redirect('configuracion_criterios')
            else:
                messages.error(request, 'Eliminación cancelada.')
                return redirect('detalle_criterio', criterio_id=criterio_id)
        
        # Mostrar confirmación
        context = {
            'criterio': criterio,
            'anomalias_asociadas': anomalias_asociadas,
            'puede_eliminar': anomalias_asociadas == 0
        }
        
        return render(request, 'anomalias/confirmar_eliminar_criterio.html', context)
        
    except Exception as e:
        print(f"❌ Error eliminando criterio: {str(e)}")
        messages.error(request, f'Error eliminando criterio: {str(e)}')
        return redirect('configuracion_criterios')

@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa', 'coordinador_carrera', 'admin'])
def listado_anomalias(request):
    """
    Lista paginada de anomalías con filtros

    🎓 APRENDIZAJE: Función basada en vista (convertida desde ListView)
    - Paginación manual
    - Ordenamiento
    - Filtros personalizados
    """
    print(f"\n📊 listado_anomalias - Usuario: {request.user.username} ({request.user.rol})")

    # ================================================================
    # QUERYSET BASE CON OPTIMIZACIÓN
    # ================================================================
    queryset = DeteccionAnomalia.objects.select_related(
        'estudiante', 'estudiante__carrera', 'criterio_usado', 'revisado_por'
    ).order_by('-fecha_deteccion')

    # ================================================================
    # FILTRAR POR ROL DEL USUARIO
    # ================================================================
    if request.user.rol == 'coordinador_carrera':
        try:
            carrera = Carrera.objects.get(coordinador=request.user)
            queryset = queryset.filter(estudiante__carrera=carrera)
            print(f"👨‍🎓 Filtrando por carrera: {carrera.nombre}")
        except Carrera.DoesNotExist:
            print("❌ Coordinador sin carrera asignada")
            queryset = queryset.none()

    # ================================================================
    # APLICAR FILTROS DE BÚSQUEDA
    # ================================================================

    # 1. Filtro por estado
    estado = request.GET.get('estado')
    if estado:
        queryset = queryset.filter(estado=estado)
        print(f"🔍 Filtro estado: {estado}")

    # 2. Filtro por tipo de anomalía
    tipo = request.GET.get('tipo')
    if tipo:
        queryset = queryset.filter(tipo_anomalia=tipo)
        print(f"🔍 Filtro tipo: {tipo}")

    # 3. Filtro por prioridad
    prioridad = request.GET.get('prioridad')
    if prioridad:
        try:
            prioridad_int = int(prioridad)
            queryset = queryset.filter(prioridad=prioridad_int)
            print(f"🔍 Filtro prioridad: {prioridad_int}")
        except ValueError:
            pass

    # 4. Filtro por carrera (para coordinadores CPA)
    carrera_filtro = request.GET.get('carrera')
    if carrera_filtro and request.user.rol in ['coordinador_cpa', 'analista_cpa']:
        try:
            carrera_obj = Carrera.objects.get(id=carrera_filtro)
            queryset = queryset.filter(estudiante__carrera=carrera_obj)
            print(f"🔍 Filtro carrera: {carrera_obj.nombre}")
        except Carrera.DoesNotExist:
            pass

    # 5. Filtro por rango de fechas
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_deteccion__date__gte=fecha_desde_obj)
            print(f"🔍 Filtro fecha desde: {fecha_desde}")
        except ValueError:
            print(f"❌ Fecha desde inválida: {fecha_desde}")

    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            queryset = queryset.filter(fecha_deteccion__date__lte=fecha_hasta_obj)
            print(f"🔍 Filtro fecha hasta: {fecha_hasta}")
        except ValueError:
            print(f"❌ Fecha hasta inválida: {fecha_hasta}")

    # 6. Filtro por nombre de estudiante
    buscar = request.GET.get('buscar')
    if buscar:
        queryset = queryset.filter(
            Q(estudiante__nombre__icontains=buscar) |
            Q(estudiante__id_estudiante__icontains=buscar)
        )
        print(f"🔍 Búsqueda: {buscar}")

    # 7. Ordenamiento
    orden = request.GET.get('orden', '-fecha_deteccion')
    if orden in ['-fecha_deteccion', 'fecha_deteccion', '-score_anomalia', 'score_anomalia',
                    'estudiante__nombre', '-estudiante__nombre', '-prioridad', 'prioridad']:
        queryset = queryset.order_by(orden)
        print(f"📋 Ordenamiento: {orden}")

    print(f"📊 Total anomalías después de filtros: {queryset.count()}")

    # ================================================================
    # PAGINACIÓN DINÁMICA
    # ================================================================
    per_page = request.GET.get('per_page', '20')
    try:
        per_page = int(per_page)
        # Limitar entre 10 y 100 elementos por página
        if not (10 <= per_page <= 100):
            per_page = 20
    except (ValueError, TypeError):
        per_page = 20

    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ================================================================
    # PREPARAR CONTEXTO
    # ================================================================

    # Obtener parámetros actuales para mantener filtros en paginación
    filtros_actuales = {
        'estado': request.GET.get('estado', ''),
        'tipo': request.GET.get('tipo', ''),
        'prioridad': request.GET.get('prioridad', ''),
        'carrera': request.GET.get('carrera', ''),
        'fecha_desde': request.GET.get('fecha_desde', ''),
        'fecha_hasta': request.GET.get('fecha_hasta', ''),
        'buscar': request.GET.get('buscar', ''),
        'orden': request.GET.get('orden', '-fecha_deteccion'),
        'per_page': request.GET.get('per_page', '20')
    }

    # Opciones para los filtros
    estados_choices = DeteccionAnomalia.ESTADOS
    tipos_choices = DeteccionAnomalia.TIPOS_ANOMALIA

    # Carreras disponibles (solo para coordinadores CPA)
    carreras_disponibles = []
    if request.user.rol in ['coordinador_cpa', 'analista_cpa']:
        carreras_disponibles = Carrera.objects.all().order_by('nombre')

    # Estadísticas rápidas
    total_anomalias = queryset.count()

    # Instancias de apoyo para derivaciones masivas
    instancias_apoyo = InstanciaApoyo.objects.filter(activo=True).order_by('nombre') 

    # Contexto completo
    context = {
        'anomalias': page_obj,  # Django espera 'object_list' o el nombre personalizado
        'page_obj': page_obj,
        'paginator': paginator,
        'is_paginated': page_obj.has_other_pages(),
        'filtros_actuales': filtros_actuales,
        'estados_choices': estados_choices,
        'tipos_choices': tipos_choices,
        'carreras_disponibles': carreras_disponibles,
        'total_anomalias': total_anomalias,
        'usuario_rol': request.user.rol,
        'instancias_apoyo': instancias_apoyo,
    }

    print(f"📋 Context data preparado - Total anomalías: {total_anomalias}")

    return render(request, 'anomalias/listado_anomalias.html', context)

@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa', 'coordinador_carrera', 'admin'])
def detalle_anomalia(request, pk):
    """
    Vista detallada de una anomalía (versión como función).
    """
    anomalia = get_object_or_404(DeteccionAnomalia, pk=pk)
    
    registros_estudiante = RegistroAcademico.objects.filter(
        estudiante=anomalia.estudiante
    ).select_related('asignatura').order_by('asignatura__semestre', 'asignatura__nombre')
    
    derivaciones = Derivacion.objects.filter(deteccion_anomalia=anomalia).select_related('instancia_apoyo', 'derivado_por').order_by('-fecha_derivacion')

    # Preparamos los datos para el gráfico
    evolucion_datos = []
    for registro in registros_estudiante:
        evolucion_datos.append({
            'asignatura': registro.asignatura.nombre,
            'semestre': registro.asignatura.semestre,
            'promedio': float(registro.promedio_notas),
            'asistencia': float(registro.porcentaje_asistencia),
            'uso_plataforma': float(registro.porcentaje_uso_plataforma)
        })
    
    estados = DeteccionAnomalia.ESTADOS

    evolucion_datos = json.dumps(evolucion_datos)

    # 4. Creamos el diccionario de 'context' manualmente
    context = {
        'anomalia': anomalia, 
        'registros_academicos': registros_estudiante,
        'derivaciones': derivaciones,
        'evolucion_datos': evolucion_datos,
        'estados': estados,
    }

    # 5. Renderizamos el 'template_name' con el 'context'
    return render(request, 'anomalias/detalle_anomalia.html', context)

@login_required
def actualizar_estado_anomalia(request, anomalia_id):
    """
    Actualiza el estado de una anomalía
    
    🎓 APRENDIZAJE: Esta función ahora es simple
    - La lógica está en el modelo
    - Solo maneja HTTP y permisos
    """
    
    anomalia = get_object_or_404(DeteccionAnomalia, id=anomalia_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        observaciones = request.POST.get('observaciones', '')
        
        try:
            # Usar método del modelo (Fat Model, Thin View)
            anomalia.actualizar_estado(nuevo_estado, observaciones, request.user)
            
            # Notificar cambio
            enviar_notificacion_cambio_estado(anomalia, nuevo_estado)
            
            messages.success(request, 'Estado actualizado correctamente')
            return render(request, 'anomalias/detalle_anomalia.html')
            
        except ValueError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return render(request, 'anomalias/detalle_anomalia.html')

@login_required
def crear_derivacion(request, anomalia_id):
    """
    Crea una derivación a instancia de apoyo
    
    🎓 APRENDIZAJE: Separación de responsabilidades
    - Vista: Maneja el formulario
    - Notificaciones: En módulo aparte
    """
    anomalia = get_object_or_404(DeteccionAnomalia, id=anomalia_id)
    
    if not anomalia.puede_ser_derivada():
        messages.error(request, 'Esta anomalía no puede ser derivada en su estado actual')
        return redirect('detalle_anomalia', pk=anomalia_id)
    
    if request.method == 'POST':
        form = DerivacionForm(request.POST)
        if form.is_valid():
            derivacion = form.save(commit=False)
            derivacion.deteccion_anomalia = anomalia
            derivacion.usuario_creador = request.user
            derivacion.save()
            
            # Actualizar estado de anomalía
            anomalia.actualizar_estado('derivada', 'Derivada a instancia de apoyo', request.user)
            
            # Notificar
            enviar_notificacion_derivacion(derivacion)
            
            messages.success(request, 'Derivación creada exitosamente')
            return redirect('detalle_anomalia', pk=anomalia_id)
    else:
        form = DerivacionForm()
    
    return render(request, 'anomalias/crear_derivacion.html', {
        'form': form,
        'anomalia': anomalia
    })

@login_required
@user_passes_test(lambda u: u.rol in ['analista_cpa', 'coordinador_cpa', 'admin'])
def gestionar_derivaciones(request):
    """Vista mejorada para gestionar derivaciones."""
    # Queryset base
    derivaciones = Derivacion.objects.select_related(
        'deteccion_anomalia__estudiante',
        'deteccion_anomalia__estudiante__carrera',
        'instancia_apoyo',
        'derivado_por'
    ).order_by('-fecha_derivacion')
    
    # Aplicar filtros
    estado = request.GET.get('estado')
    if estado:
        derivaciones = derivaciones.filter(estado=estado)
    
    instancia = request.GET.get('instancia')
    if instancia:
        derivaciones = derivaciones.filter(instancia_apoyo_id=instancia)
    
    fecha_desde = request.GET.get('fecha_desde')
    if fecha_desde:
        derivaciones = derivaciones.filter(fecha_derivacion__date__gte=fecha_desde)
    
    busqueda = request.GET.get('busqueda')
    if busqueda:
        derivaciones = derivaciones.filter(
            Q(deteccion_anomalia__estudiante__nombre__icontains=busqueda) |
            Q(deteccion_anomalia__estudiante__id_estudiante__icontains=busqueda)
        )
    
    # Estadísticas rápidas
    total_derivaciones = derivaciones.count()
    derivaciones_pendientes = derivaciones.filter(estado='pendiente').count()
    derivaciones_proceso = derivaciones.filter(estado='en_proceso').count()
    derivaciones_completadas = derivaciones.filter(estado='completada').count()
    
    # Paginación
    paginator = Paginator(derivaciones, 15)
    page = request.GET.get('page')
    derivaciones_paginadas = paginator.get_page(page)
    
    context = {
        'derivaciones': derivaciones_paginadas,
        'derivaciones_pendientes': derivaciones_pendientes,
        'derivaciones_proceso': derivaciones_proceso,
        'derivaciones_completadas': derivaciones_completadas,
        'total_derivaciones': total_derivaciones,
        'estados_derivacion': Derivacion.ESTADOS_DERIVACION,
        'instancias_apoyo': InstanciaApoyo.objects.filter(activo=True),
    }
    
    return render(request, 'anomalias/gestionar_derivaciones.html', context)













